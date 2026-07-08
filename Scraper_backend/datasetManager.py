# this code is the excel converter but now with the change that it can read both yyyy-mm-dd and dd-mm-yyyy and calculate the days remaining correctly
# this code also has the new sectors added: the corporate and pet
# please change any other code as this has been changed
# i have also added a cell which has connected the llm cell so this one with the json to excel converter so it is just once cell that can do it all but i still added them separately
import os
import json
import re
import pandas as pd
from datetime import datetime
import requests
from openpyxl.styles import PatternFill

# =====================================================================
# PART 2: LIVE NETWORK RATE FETCHER
# =====================================================================
def fetch_live_rates_or_die():
    url = "https://open.er-api.com/v6/latest/USD"
    response = requests.get(url, timeout=12)

    if response.status_code != 200:
        raise RuntimeError(f"CRITICAL: Live API down (Status {response.status_code}).")

    raw_data = response.json()
    if "rates" not in raw_data:
        raise ValueError("CRITICAL: Invalid data payload from currency server.")

    usd_rates = raw_data["rates"]
    required_symbols = ["INR", "EUR", "GBP", "AUD", "JPY", "AED", "CAD", "KRW", "SGD"]
    for symbol in required_symbols:
        if symbol not in usd_rates:
            raise ValueError(f"CRITICAL: Token '{symbol}' missing from market feed.")

    usd_to_inr = float(usd_rates["INR"])

    live_inr_matrix = {
        "USD": round(usd_to_inr, 4),
        "EUR": round(usd_to_inr / usd_rates["EUR"], 4),
        "GBP": round(usd_to_inr / usd_rates["GBP"], 4),
        "AUD": round(usd_to_inr / usd_rates["AUD"], 4),
        "JPY": round(usd_to_inr / usd_rates["JPY"], 4),
        "AED": round(usd_to_inr / usd_rates["AED"], 4),
        "CAD": round(usd_to_inr / usd_rates["CAD"], 4),
        "KRW": round(usd_to_inr / usd_rates["KRW"], 4),
        "SGD": round(usd_to_inr / usd_rates["SGD"], 4),
        "INR": 1.0
    }

    print(" 🔄 LIVE PARITY RATIOS LOCKED TO 4 DECIMAL PLACES:")
    for currency, rate in live_inr_matrix.items():
        if currency != "INR":
            print(f"   1 {currency} = {rate} INR")
    print("="*60 + "\n")

    return live_inr_matrix

# =====================================================================
# PART 3: STRING UTILITIES, DATE PARSERS & CONVERSION ENGINES
# =====================================================================
def parse_flexible_date(date_str):
    if not date_str or pd.isna(date_str) or str(date_str).strip().lower() in ["none", "", "n/a", "-"]:
        return None

    date_str = str(date_str).strip()
    
    # CRITICAL FIX: Strip ISO time elements (e.g., "2026-06-25T14:30:00Z" becomes "2026-06-25")
    # But leave spaces alone if they are separation marks for textual months (e.g., "25 Jun 2026")
    if 'T' in date_str:
        date_str = date_str.split('T')[0]

    # Expanded robust format list including 3-letter months in both ascending and descending orders
    allowed_formats = (
        # --- Pure Numeric Formats ---
        "%Y-%m-%d", "%Y/%m/%d",  # Descending numeric (e.g., 2026-06-25)
        "%d-%m-%Y", "%d/%m/%Y",  # Ascending numeric (e.g., 25-06-2026)
        
        # --- 3-Letter Month ASCENDING Formats (e.g., 25-Jun-2026, 25/Jun/2026, 25 Jun 2026) ---
        "%d-%b-%Y", "%d/%b/%Y", "%d %b %Y",
        
        # --- 3-Letter Month DESCENDING Formats (e.g., 2026-Jun-25, 2026/Jun/25, 2026 Jun 25) ---
        "%Y-%b-%d", "%Y/%b/%d", "%Y %b %d",
        
        # --- Full Written Month Name Formats ---
        "%B %d, %Y", "%d %B %Y"
    )

    for fmt in allowed_formats:
        try:
            return datetime.strptime(date_str, fmt)
        except ValueError:
            continue
    return None

def clean_tender_string(value_str):
    if not value_str or pd.isna(value_str):
        return ""
    cleaned = str(value_str).upper().strip()
    cleaned = re.sub(r'(ID|REF|NO|NUMBER|VERSION)[:.\s]*\d+', '', cleaned)
    cleaned = re.sub(r'\b(19|20)\d{2}\b', '', cleaned)
    cleaned = re.sub(r'\s+', ' ', cleaned)
    return cleaned

def extract_clean_float(cleaned_str):
    num_match = re.search(r'([\d.,]+)', cleaned_str)
    if not num_match:
        return None

    num_str = num_match.group(1)

    if "." in num_str and "," in num_str:
        if num_str.find(".") < num_str.find(","):
            num_str = num_str.replace(".", "").replace(",", ".")
        else:
            num_str = num_str.replace(",", "")
    elif "," in num_str:
        parts = num_str.split(",")
        if len(parts) == 2 and len(parts[1]) <= 2:
            num_str = num_str.replace(",", ".")
        else:
            num_str = num_str.replace(",", "")
    elif "." in num_str:
        parts = num_str.split(".")
        if len(parts) == 2 and len(parts[1]) == 3 and not cleaned_str.endswith(parts[1]):
            num_str = num_str.replace(".", "")

    try:
        return float(num_str)
    except ValueError:
        return None

def get_verbal_scale_multiplier(cleaned_str):
    if re.search(r'\d\s*M\b', cleaned_str) or "MILLION" in cleaned_str or "MIO" in cleaned_str or "88.5M" in cleaned_str.replace(" ", ""):
        return 1_000_000.0
    if "BILLION" in cleaned_str or re.search(r'\bB\b', cleaned_str):
        return 1_000_000_000.0
    if "LAKH" in cleaned_str or "LC" in cleaned_str:
        return 100_000.0
    if "CRORE" in cleaned_str or "CR" in cleaned_str:
        return 10_000_000.0
    if "K" in cleaned_str:
        return 1_000.0
    return 1.0

def identify_currency_type(cleaned_str):
    if "C$" in cleaned_str: return "CAD"
    if "S$" in cleaned_str: return "SGD"
    if "USD" in cleaned_str or "$" in cleaned_str: return "USD"
    if "EUR" in cleaned_str or "€" in cleaned_str: return "EUR"
    if "GBP" in cleaned_str or "£" in cleaned_str: return "GBP"
    if "AUD" in cleaned_str: return "AUD"
    if "JPY" in cleaned_str or "¥" in cleaned_str: return "JPY"
    if "AED" in cleaned_str or "DIRHAM" in cleaned_str or "DH" in cleaned_str: return "AED"
    if "CAD" in cleaned_str: return "CAD"
    if "KRW" in cleaned_str or "WON" in cleaned_str or "₩" in cleaned_str: return "KRW"
    if "SGD" in cleaned_str: return "SGD"
    return "INR"

def parse_and_convert_to_inr(value_str, assigned_currency_code, live_rates):
    if not value_str or pd.isna(value_str):
        return "N/A"

    cleaned_text = clean_tender_string(value_str)
    base_number = extract_clean_float(cleaned_text)

    if base_number == 0.0 or base_number is None:
        return value_str

    scale_factor = get_verbal_scale_multiplier(cleaned_text)
    true_base_value = round(base_number * scale_factor, 2)
    final_inr = true_base_value * live_rates.get(assigned_currency_code, 1.0)

    return f"{final_inr:,.2f} INR"

def calculate_score_color(score):
    try:
        val = max(0.0, min(10.0, float(score)))
    except (ValueError, TypeError):
        return "FFFFFF"

    if val < 5.0:
        ratio = val / 5.0
        r = 255
        g = int(77 + (138 * ratio))
        b = int(77 - (77 * ratio))
    else:
        ratio = (val - 5.0) / 5.0
        r = int(255 - (179 * ratio))
        g = int(215 - (40 * ratio))
        b = int(0 + (80 * ratio))

    return f"{r:02X}{g:02X}{b:02X}"

# =====================================================================
# PART 4: EXCEL GENERATION PIPELINE 
# =====================================================================
def json_to_excel(json_filename="all_tenders.json", excel_filename="all_tenders_pipeline.xlsx"):
    try:
        if os.path.exists(excel_filename):
            try:
                os.remove(excel_filename)
            except OSError:
                pass

        with open(json_filename, "r") as f:
            data = json.load(f)

        if not data:
            print("⚠️ Operations halted: Target JSON data file is empty.")
            return

        live_exchange_rates = fetch_live_rates_or_die()
        excel_rows = []
        current_year = datetime.now().year

        for index, tender in enumerate(data, start=1):
            primary_key = f"TND-{current_year}-{index:04d}"

            min_val_str = tender.get("Budget in Local Currency Minimum", "")
            max_val_str = tender.get("Budget in Local Currency Maximum", "")
            
            # explicit_currency = tender.get("Budget Currency", "").strip().upper()
            explicit_currency = str(tender.get("Budget Currency") or "").strip().upper()
            
            if explicit_currency and explicit_currency in live_exchange_rates:
                currency_code = explicit_currency
            else:
                cleaned_budget_str = clean_tender_string(str(min_val_str))
                currency_code = identify_currency_type(cleaned_budget_str)
            
            inr_min_value = parse_and_convert_to_inr(min_val_str, currency_code, live_exchange_rates)
            inr_max_value = parse_and_convert_to_inr(max_val_str, currency_code, live_exchange_rates)

            opening_date_raw = tender.get("Opening Date") if tender.get("Opening Date") else tender.get("BidOpeningDate", "")
            closing_date_raw = tender.get("Expiry Date") if tender.get("Expiry Date") else tender.get("BidEndDate", "")

            parsed_open_dt = parse_flexible_date(opening_date_raw)
            parsed_close_dt = parse_flexible_date(closing_date_raw)

            opening_date_str = parsed_open_dt.strftime("%Y-%m-%d") if parsed_open_dt else str(opening_date_raw)
            closing_date_str = parsed_close_dt.strftime("%Y-%m-%d") if parsed_close_dt else str(closing_date_raw)

            days_remaining = "N/A"
            if parsed_close_dt:
                delta = parsed_close_dt.date() - datetime.now().date()
                days_remaining = max(0, delta.days)

            # Derive status from the closing date instead of hardcoding "Open":
            #   - no / unparseable closing date -> Unknown
            #   - closing date already passed    -> Closed
            #   - closing date today or future   -> Open
            if parsed_close_dt is None:
                resolved_status = "Unknown"
            elif parsed_close_dt.date() < datetime.now().date():
                resolved_status = "Closed"
            else:
                resolved_status = "Open"

            sector_value = tender.get("Sector")
            if not sector_value or sector_value == "None":
                sector_value = tender.get("Health or Defence Category", "N/A")

            row = {
                "Primary Key": primary_key,
                "Relevancy Score": float(tender.get("LLM_RelevancyScore", 0.0)),
                "Tender Title": tender.get("Tender Title", ""),
                "Description": tender.get("Tender Description", ""),
                "Organisation name": tender.get("Organisation Name", ""),
                "Tender URL": tender.get("Link to the Tender", ""),
                "Original Currency": currency_code,
                "Original Currency Minimum": min_val_str,
                "Original Currency Maximum": max_val_str,
                "INR Budget Minimum": inr_min_value,
                "INR Budget Maximum": inr_max_value,
                "Sector": sector_value,
                "Opening date": opening_date_str,
                "Closing date": closing_date_str,
                "Days remaining": days_remaining,
                "Tender Status": resolved_status,
                "Award Date": tender.get("Award Date") if tender.get("Award Date") else tender.get("AwardDate", "N/A"),
                "Country": tender.get("Country", "")
            }
            excel_rows.append(row)

        df = pd.DataFrame(excel_rows)

        columns_order = [
            "Primary Key", "Relevancy Score", "Tender Title", "Description",
            "Organisation name", "Tender URL", "Original Currency", "Original Currency Minimum", "Original Currency Maximum",
            "INR Budget Minimum", "INR Budget Maximum", "Sector", "Opening date", "Closing date",
            "Days remaining", "Tender Status", "Award Date", "Country"
        ]
        df = df[columns_order]

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

        with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name="Tenders")
            worksheet = writer.sheets["Tenders"]

            score_col_idx = columns_order.index("Relevancy Score") + 1
            status_col_idx = columns_order.index("Tender Status") + 1

            for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                score_cell = worksheet.cell(row=row_idx, column=score_col_idx)
                status_cell = worksheet.cell(row=row_idx, column=status_col_idx)

                hex_color = calculate_score_color(score_cell.value)
                score_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

                if status_cell.value == "Open":
                    status_cell.fill = green_fill
                elif status_cell.value == "Coming Soon":
                    status_cell.fill = yellow_fill
                elif status_cell.value == "Closed":
                    status_cell.fill = red_fill
                elif status_cell.value == "Unknown":
                    status_cell.fill = gray_fill

            for col in worksheet.columns:
                max_len = 0
                col_letter = col[0].column_letter
                for cell in col:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                worksheet.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 60)

        print(f"✨ Success! Saved precise updates to '{excel_filename}'.")

    except Exception as e:
        print(f"\n❌ CRITICAL EXCEL PIPELINE FAILURE: {e}\n")
        raise

if __name__ == "__main__":
    print("📊 Initializing automatic Excel Compilation Pipeline...")
    json_to_excel(
        json_filename="all_tenders.json",
        excel_filename="all_tenders_pipeline.xlsx"
    )
