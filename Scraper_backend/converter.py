import json
from openpyxl import Workbook

# -----------------------------
# Input and Output files
# -----------------------------
INPUT_JSON = "all_tenders.json"
OUTPUT_EXCEL = "all_tenders.xlsx"

# -----------------------------
# Excel column headers
# -----------------------------
headers = [
    "Primary Key",
    "Relevancy Score",
    "Tender Title",
    "Description",
    "Organisation name",
    "Tender URL",
    "Original Currency",
    "Original Currency Minimum",
    "Original Currency Maximum",
    "INR Budget Minimum",
    "INR Budget Maximum",
    "Sector",
    "Opening date",
    "Closing date",
    "Days remaining",
    "Tender Status",
    "Award Date",
    "Country"
]

# -----------------------------
# Load JSON
# -----------------------------
with open(INPUT_JSON, "r", encoding="utf-8") as f:
    data = json.load(f)

# -----------------------------
# Create workbook
# -----------------------------
wb = Workbook()
ws = wb.active
ws.title = "Tenders"

# Write headers
ws.append(headers)

# -----------------------------
# Write data
# -----------------------------
for item in data:

    # Calculate relevancy score
    scores = [
        item.get("health_score"),
        item.get("defence_score"),
        item.get("corporate_score"),
        item.get("pet_score")
    ]

    scores = [s for s in scores if isinstance(s, (int, float))]
    relevancy_score = max(scores) if scores else None

    ws.append([
        item.get("Primary Key"),
        relevancy_score,
        item.get("Tender Title"),
        item.get("Tender Description"),
        item.get("Organisation Name"),
        item.get("Link to the Tender"),
        item.get("Budget Currency"),
        item.get("Budget in Local Currency Minimum"),
        item.get("Budget in Local Currency Maximum"),
        item.get("Budget in INR Minimum"),
        item.get("Budget in INR Maximum"),
        item.get("Sector"),
        item.get("Opening Date"),
        item.get("Expiry Date"),
        item.get("Timeline"),
        item.get("Tender Status"),
        item.get("Award Date"),
        item.get("Country")
    ])

# -----------------------------
# Save workbook
# -----------------------------

from openpyxl.utils import get_column_letter

# Auto-adjust column widths
for column_cells in ws.columns:
    max_length = 0
    column = get_column_letter(column_cells[0].column)

    for cell in column_cells:
        try:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))
        except Exception:
            pass

    # Add a little padding
    adjusted_width = max_length + 2

    # Optional: limit excessively wide columns
    adjusted_width = min(adjusted_width, 100)

    ws.column_dimensions[column].width = adjusted_width
    
wb.save(OUTPUT_EXCEL)

print(f"Excel file saved as: {OUTPUT_EXCEL}")